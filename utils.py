import csv
from dataclasses import dataclass
from openpyxl import load_workbook, Workbook
from pathlib import Path
import sys
import typing
from typing import List, Union

from gtts import gTTS


@dataclass
class Exercise:
    """Class for holding the details of an exercise segment"""
    name: str
    duration: int  # seconds
    reps: typing.Optional[int] = None
    language: str = "en"
    slow: bool = False

    def _instruction_text(self) -> str:
        if self.reps:
            text = f"{self.reps} {self.name} in {self.duration} seconds"
        else:
            text = f"{self.name} for {self.duration} seconds"
        return text

    def create_speech_obj(self) -> gTTS:
        speech_obj = gTTS(text=self._instruction_text(), lang=self.language, slow=self.slow)
        return speech_obj


def load_exercises_from_xlsx(filepath_or_workbook: Union[Path, Workbook]) -> List[Exercise]:
    """
    Loads Exercise objects from an xlsx file
    """
    if type(filepath_or_workbook) == Workbook:
        wb = filepath_or_workbook
    elif type(filepath_or_workbook) == Path:
        wb = load_workbook(filepath_or_workbook)
    ws = wb.active
    exercises = []
    for i, row in enumerate(ws.iter_rows()):
        if i == 0:
            continue
        if row[2].value:
            exercise = Exercise(row[0].value, int(row[1].value), int(row[2].value))
        else:
            exercise = Exercise(row[0].value, int(row[1].value))
        exercises.append(exercise)
    return exercises


def load_exercises_from_csv(filepath: Path) -> List[Exercise]:
    """
    Converts a CSV into a Workbook object to be processed into exercises just like an xlsx would.
    """
    wb = Workbook()
    ws = wb.active

    with open(filepath, "rt", encoding="utf-8") as f:
        reader = csv.reader(f)
        for row in reader:
            ws.append(row)

    exercises = load_exercises_from_xlsx(wb)
    return exercises


def load_exercises_from_stdin() -> List[Exercise]:
    """
    Loads Exercise objects from a comma-separated list sent from stdin
    """
    rows = [e.strip() for e in sys.stdin]
    exercises = []
    for row in rows:
        items = row.split(",")
        if len(items) == 3:
            exercise = Exercise(items[0], int(items[1]), int(items[2]))
        else:
            exercise = Exercise(items[0], int(items[1]))
        exercises.append(exercise)
    return exercises
