import argparse
from collections import namedtuple
from dataclasses import dataclass
from pathlib import Path
import typing
import os
import sys

from gtts import gTTS
from gtts.tts import gTTSError
from pydub import AudioSegment


@dataclass
class Exercise:
    """Class for holding the details of an exercise segment"""
    name: str
    duration: int  # seconds
    reps: typing.Optional[int] = None
    language: str = "en"
    slow: bool = False

    def _instruction_text(self) -> str:
        if self.reps:
            text = f"{self.reps} {self.name} in {self.duration} seconds"
        else:
            text = f"{self.name} for {self.duration} seconds"
        return text

    def create_speech_obj(self) -> gTTS:
        speech_obj = gTTS(text=self._instruction_text(), lang=self.language, slow=self.slow)
        return speech_obj


class Mp3Creator:
    """Joins Exercise objects along with beeps and silence intervals into a single file"""

    Segment = namedtuple('Segment', ['exercise', 'speech_obj', 'audio_file_path'])

    def __init__(self, exercises: list):
        self.segments = self._create_segments(exercises)
        self._load_audio_segments_from_resources()

    def _create_segments(self, exercises: list):
        segments = []
        for i, exercise in enumerate(exercises):
            speech_obj = exercise.create_speech_obj()
            filepath = Path(".") / f"{exercise.name[:20]}.mp3"
            segment = self.Segment(exercise, speech_obj, filepath)
            segments.append(segment)
        return segments

    def _load_audio_segments_from_resources(self):
        p = Path(__file__).parent / "audio_resources"
        self.beep_start = AudioSegment.from_mp3(p / "beep_start.mp3")
        self.beep_intermediate = AudioSegment.from_mp3(p / "beep_intermediate.mp3")
        self.beep_end = AudioSegment.from_mp3(p / "beep_end.mp3")
        self.finished_sound = AudioSegment.from_mp3(p / "workout_end.mp3")

        self.pause = AudioSegment.silent(2500)  # milliseconds

    def create_mp3s(self, tags: dict = None):
        self._create_segment_files(tags)

    def merge_mp3s_into_single_file(self, filepath: Path, tags: dict = None):

        final_mp3 = AudioSegment.silent(2000)
        for segment in self.segments:
            final_mp3 += AudioSegment.from_mp3(segment.audio_file_path)
        final_mp3 += self.finished_sound

        final_mp3.export(filepath, format="mp3", tags=tags)

        self._delete_segment_files()

        return filepath

    def _create_segment_files(self, tags: dict = None):
        counter = 0
        total_tracks = len(self.segments)
        for segment in self.segments:
            while True:
                try:
                    segment.speech_obj.save(segment.audio_file_path)
                except gTTSError:
                    print("Error downloading segment file.")
                break
            file = self._add_beeps_and_silence_to_segment_file(segment)
            counter += 1
            track_tags = {
                "artist": tags.get("artist", "Unknown Artist"),
                "album": tags.get("album", "Unknown Album"),
                "title": segment.exercise.name,
                "track": f"{counter}/{total_tracks}"
            }
            file.export(segment.audio_file_path, format="mp3", tags=track_tags)

    def _add_beeps_and_silence_to_segment_file(self, segment: Segment) -> Segment:
        speech = AudioSegment.from_mp3(segment.audio_file_path)
        AS_MILLISECONDS = 1000
        BEEP_INTERVAL = 5 * AS_MILLISECONDS
        silence = AudioSegment.silent(segment.exercise.duration * AS_MILLISECONDS)
        for interval in range(BEEP_INTERVAL, len(silence), BEEP_INTERVAL):
            silence = silence[:interval] + \
                      self.beep_intermediate + \
                      silence[interval + len(self.beep_intermediate):]
        silence_with_beeps = self.beep_start + silence[len(self.beep_start):] + self.beep_end
        new_audio = speech + self.pause + silence_with_beeps

        return new_audio

    def _delete_segment_files(self):
        for segment in self.segments:
            os.remove(segment.audio_file_path)


def load_exercises_from_xlsx(filepath: Path):
    from openpyxl import load_workbook

    wb = load_workbook(filepath)
    ws = wb.active
    exercises = []
    for i, row in enumerate(ws.iter_rows()):
        if i == 0:
            continue
        if row[2].value:
            exercise = Exercise(row[0].value, int(row[1].value), int(row[2].value))
        else:
            exercise = Exercise(row[0].value, int(row[1].value))
        exercises.append(exercise)
    return exercises


def load_exercises_from_stdin():
    rows = [e.strip() for e in sys.stdin]
    exercises = []
    for row in rows:
        items = row.split(",")
        if len(items) == 3:
            exercise = Exercise(items[0], int(items[1]), int(items[2]))
        else:
            exercise = Exercise(items[0], int(items[1]))
        exercises.append(exercise)
    return exercises


if __name__ == '__main__':

    parser = argparse.ArgumentParser(
        description="Create audio file from a list of text elements, each with a specified silence interval to follow.")
    parser.add_argument("output", type=lambda p: Path(p).absolute(),
                        help="Name of the output file, optionally include full filepath")

    parser.add_argument("--title", type=str,
                        help="Name of 'title' to help with classification of file. Uses filename if not provided.")
    parser.add_argument("--artist", type=str, help="Name of 'artist' to help with classification of file.")
    parser.add_argument("--album", type=str, help="Name of 'album' to help with classification of file.")

    parser.add_argument("--merge", action='store_true', help="Merge audio files into a single audio file.")

    group = parser.add_mutually_exclusive_group()
    group.add_argument("--xlsx", type=lambda p: Path(p).absolute(), help="Import from an xlsx file.")
    group.add_argument("--csv", type=lambda p: Path(p).absolute(), help="Import from a CSV file.")
    group.add_argument("--stdin", action="store_true",
                       help="Import from stdin (comma-separated with each exercise on a new line).")

    args = parser.parse_args()

    if args.output.suffix != ".mp3":
        args.output = args.output.parent / (args.output.stem + '.mp3')

    tags = {}
    if args.title:
        tags["title"] = args.title
    else:
        tags["title"] = args.output.stem
    if args.artist:
        tags["artist"] = args.artist
    if args.album:
        tags["album"] = args.album

    if args.xlsx:
        exercises = load_exercises_from_xlsx(args.xlsx)
    if args.csv:
        pass
    if args.stdin:
        exercises = load_exercises_from_stdin()

    creator = Mp3Creator(exercises)
    creator.create_mp3s(tags)
    if args.merge:
        creator.merge_mp3s_into_single_file(args.output, tags)
